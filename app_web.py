from flask import Flask, render_template, request, redirect, send_file, session
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
import sqlite3
import datetime
import os


# USUARIO = "admin"
# PASSWORD = "1234"

app = Flask(__name__)


app.secret_key = "clave_secreta"


#BASE DE DATOS
# Crear base de datos
def crear_base():
    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()

    # Crear Tabla calibraciones
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS calibraciones_4 (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        instrumento TEXT,
        tipo TEXT,
        fecha TEXT,
        frecuencia INTEGER
    )
    """)

    #  Crear tabla usuarios
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario TEXT UNIQUE,
        password TEXT,
        rol TEXT
    )
    """)

    conexion.commit()
    conexion.close()


# Crear usuario inicial

def crear_usuario_inicial():
    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()

    try:
        password_hash = generate_password_hash("1234")

        cursor.execute(
            "INSERT INTO usuarios (usuario, password, rol) VALUES (?, ?, ?)",
            ("admin", password_hash, "admin")
        )

        conexion.commit()
    except:
        pass

    conexion.close()

crear_base()
crear_usuario_inicial()

# Función para parsear fechas con diferentes formatos
def parse_fecha(fecha_str):
    for formato in ("%Y-%m-%d", "%Y_%m_%d"):
        try:
            return datetime.datetime.strptime(fecha_str, formato).date()
        except ValueError:
            continue
    return None

# Limpiar formato de fechas antiguas
conexion = sqlite3.connect("calibraciones_4.db")
cursor = conexion.cursor()

cursor.execute("SELECT id, fecha FROM calibraciones_4")
datos = cursor.fetchall()

for id, fecha in datos:
    fecha_limpia = fecha.replace("_", "-")
    cursor.execute(
        "UPDATE calibraciones_4 SET fecha=? WHERE id=?",
        (fecha_limpia, id)
    )

conexion.commit()
conexion.close()


# RUTAS

# Ruta para admin
@app.route("/admin")
def admin():
    if not session.get("logueado"):
        return redirect("/login")

    if session.get("rol") != "admin":
        return "⛔ Acceso solo para administradores"

    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()

    cursor.execute("SELECT id, usuario, rol FROM usuarios")
    usuarios = cursor.fetchall()

    conexion.close()

    return render_template("admin.html", usuarios=usuarios)

# Ruta para crear nuevo usuario
@app.route("/crear_usuario", methods=["POST"])
def crear_usuario():
    if session.get("rol") != "admin":
        return "⛔ Sin permisos"

    usuario = request.form["usuario"]
    password = request.form["password"]
    rol = request.form["rol"]

    from werkzeug.security import generate_password_hash
    password_hash = generate_password_hash(password)

    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()

    try:
        cursor.execute(
            "INSERT INTO usuarios (usuario, password, rol) VALUES (?, ?, ?)",
            (usuario, password_hash, rol)
        )
        conexion.commit()
    except:
        return "⚠️ Usuario ya existe"

    conexion.close()

    return redirect("/admin")


# Ruta para eliminar usuario
# @app.route("/eliminar_usuario/<int:id>")
# def eliminar_usuario(id):
#     if session.get("rol") != "admin":
#         return "⛔ Sin permisos"
    
    

#     conexion = sqlite3.connect("calibraciones_4.db")
#     cursor = conexion.cursor()

    

#     cursor.execute("DELETE FROM usuarios WHERE id=?", (id,))
#     conexion.commit()
#     conexion.close()

#     return redirect("/admin")

@app.route("/eliminar_usuario/<int:id>")
def eliminar_usuario(id):
    if not session.get("logueado"):
        return redirect("/login")

    if session.get("rol") != "admin":
        return "⛔ Sin permisos"

    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()

    # 🔍 Buscar usuario a eliminar
    cursor.execute("SELECT usuario FROM usuarios WHERE id=?", (id,))
    resultado = cursor.fetchone()

    if resultado:
        usuario_a_eliminar = resultado[0]

        # 🚫 Evitar eliminarse a sí mismo
        if session.get("usuario") == usuario_a_eliminar:
            conexion.close()
            return "⛔ No podés eliminar tu propio usuario"

        # ✅ Eliminar
        cursor.execute("DELETE FROM usuarios WHERE id=?", (id,))
        conexion.commit()

    conexion.close()

    return redirect("/admin")

#Index: muestra tabla con estado de calibraciones
@app.route("/")
def index():
    if not session.get("logueado"):
        return redirect("/login")

    
    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()

    cursor.execute("SELECT instrumento, tipo, fecha, frecuencia FROM calibraciones_4")
    datos_db = cursor.fetchall()

    conexion.close()

    datos = []
    hoy = datetime.date.today()

    for instrumento, tipo, fecha, frecuencia in datos_db:
        # fecha_cal = datetime.datetime.strptime(fecha, "%Y-%m-%d").date()
        fecha_cal = parse_fecha(fecha)

        if not fecha_cal:
            continue  # salta registros rotos

        proxima = fecha_cal + datetime.timedelta(days=frecuencia)

        dias_restantes = (proxima - hoy).days

        if hoy >= proxima:
            estado = "vencido"
        elif dias_restantes <= 7:
            estado = "proximo"
        else:
            estado = "ok"

        datos.append({
            "instrumento": instrumento,
            "tipo": tipo,
            "fecha": fecha,
            "frecuencia": frecuencia,
            "proxima": proxima,
            "estado": estado
        })

    return render_template("index.html", datos=datos)

# Ruta para dashboard con estadísticas
@app.route("/dashboard")
def dashboard():
    if not session.get("logueado"):
        return redirect("/login")

    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()

    cursor.execute("SELECT fecha, frecuencia FROM calibraciones")
    datos_db = cursor.fetchall()

    conexion.close()

    hoy = datetime.date.today()

    vencidos = 0
    proximos = 0
    ok = 0

    for fecha, frecuencia in datos_db:
        # fecha_cal = datetime.datetime.strptime(fecha, "%Y_%m_%d").date()
        fecha_cal = parse_fecha(fecha)

        if not fecha_cal:
            continue  # evita que rompa todo

        proxima = fecha_cal + datetime.timedelta(days=frecuencia)

        if hoy >= proxima:
            vencidos += 1
        elif (proxima - hoy).days <= 7:
            proximos += 1
        else:
            ok += 1

    return render_template("dashboard.html",
                           vencidos=vencidos,
                           proximos=proximos,
                           ok=ok)

# Ruta para agregar nueva calibración
@app.route("/agregar", methods=["POST"])
def agregar():
    if not session.get("logueado"):
        return redirect("/login")
    

    instrumento = request.form["instrumento"]
    tipo = request.form["tipo"]
    frecuencia = int(request.form["frecuencia"])

    # fecha = datetime.date.today()
    fecha = datetime.date.today().strftime("%Y-%m-%d")

    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()

    cursor.execute(
        "INSERT INTO calibraciones_4 (instrumento, tipo, fecha, frecuencia) VALUES (?, ?, ?, ?)",
        (instrumento, tipo, fecha, frecuencia)
    )

    conexion.commit()
    conexion.close()

    return redirect("/")


# Ruta para exportar a Excel
@app.route("/exportar")
def exportar():
    if not session.get("logueado"):
        return redirect("/login")
    
    if session.get("rol") != "admin":
        return "⛔ No tenés permisos"
    
    conexion = sqlite3.connect("calibraciones_4.db")
    cursor = conexion.cursor()


    cursor.execute("""
    SELECT instrumento, tipo, fecha, frecuencia
    FROM calibraciones_4
    """)

    datos_db = cursor.fetchall()
    conexion.close()


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calibraciones"

    # Encabezados
    
    headers = ["Instrumento", "Tipo", "Última", "Próxima", "Estado"]
    ws.append(headers)

    for col in ws[1]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal="center")

    hoy = datetime.date.today()

    
    for instrumento, tipo, fecha, frecuencia in datos_db:
        # fecha_cal = datetime.datetime.strptime(fecha, "%Y-%m-%d").date()
        fecha_cal = parse_fecha(fecha)

        if not fecha_cal:
            continue  # salta registros rotos
        proxima = fecha_cal + datetime.timedelta(days=frecuencia)

        if hoy >= proxima:
            estado = "VENCIDO"
            color = "FFCCCC"  # rojo
        elif (proxima - hoy).days <= 7:
            estado = "PRÓXIMO"
            color = "FFF3CD"  # amarillo
        else:
            estado = "OK"
            color = "CCFFCC"  # verde

        fila = [instrumento, tipo, str(fecha_cal), str(proxima), estado]
        ws.append(fila)

        fila_excel = ws.max_row

        for col in range(1, 6):
            ws.cell(row=fila_excel, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


    columnas = ["A", "B", "C", "D", "E"]

    for col in columnas:
        ws.column_dimensions[col].width = 20

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    archivo = "reporte_calibraciones.xlsx"
    wb.save(archivo)

    return send_file(archivo, as_attachment=True)

# Ruta para login
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form["usuario"]
        password = request.form["password"]

        conexion = sqlite3.connect("calibraciones_4.db")
        cursor = conexion.cursor()
        
        cursor.execute(
            "SELECT password, rol FROM usuarios WHERE usuario=?",
            (user,)
        )

        resultado = cursor.fetchone()

        if resultado:
            password_hash, rol = resultado

            if check_password_hash(password_hash, password):
                session["logueado"] = True
                session["usuario"] = user
                session["rol"] = rol
                return redirect("/")
        
        return "❌ Usuario o contraseña incorrectos"

    # return render_template("login.html")
    return render_template("login.html", error="Usuario o contraseña incorrectos")

# Ruta para logout
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")



# Iniciar la aplicación
if __name__ == "__main__":
    import traceback

    try:
        port = int(os.environ.get("PORT", 5000))
        app.run(host="0.0.0.0", port=port)
    except Exception as e:
        print("ERROR AL INICIAR:")
        traceback.print_exc()

# if __name__ == "__main__":
#     app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))